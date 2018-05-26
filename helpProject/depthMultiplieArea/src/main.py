#!/usr/bin/env python
# -*- encoding:utf-8 -*-

import logging
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

logging.basicConfig(level=logging.DEBUG)
debug = logging.debug

__author__ = 'LY'

'''
    使用openpyxl库处理
    18-5-25
'''


def task():
    r"""
    从excel获取处理数据,返回处理结果
    :return:
    """
    file_name = "../res/6Element.xlsx"
    wb = load_workbook(filename=file_name)
    # 显示所有表明
    debug(r"../res/6Element.xlsx中表名: %s", wb.get_sheet_names())
    # 获取表名为面积的表
    ws = wb["面积"]
    columns = ws.columns

    # 获取总行数
    max_rows = ws.max_row
    debug("总行数: %s", max_rows)

    # 获取面积数据
    area_dict = dict()
    for index in range(2, max_rows):  # 数据索引从1开始
        area_num = ws.cell(row=index, column=1).value
        area = ws.cell(row=index, column=2).value
        if area_num and area:
            area_dict[area_num] = area
    debug("共%s条有效面积数据", len(area_dict))

    # 获取水深数据
    depth_dict = dict()
    for index in range(2, max_rows):
        depth_num = ws.cell(row=index, column=5).value
        depth = ws.cell(row=index, column=6).value
        if depth_num and depth:
            depth_dict[depth_num] = depth
    debug("共%s条有效水深数据", len(depth_dict))

    # 整理数据
    result_list = []
    try:
        for area_num in area_dict:
            print(area_num)
            if area_num in depth_dict:
                print(area_dict[area_num], depth_dict[area_num])
                result_list.append([area_num, area_dict[area_num] * depth_dict[area_num]])
        debug("共%s条面积和水深相同数据", len(result_list))
    except IndexError as e:
        debug("area_num: %s", area_num)

    return result_list


def write2sheet(data):
    r"""
    写到表result.xlsx里去
    :param data:
    :return:
    """
    try:
        dest_file = "../res/result.xlsx"
        wb = Workbook()
        res_ws = wb.active
        res_ws.title = "体积"
        res_ws.append(["编号", "面积乘水深"])
        for item in data:
            res_ws.append(item)
        wb.save(dest_file)
    except IOError as ioErr:
        debug(ioErr)


def main():
    data = task()
    write2sheet(data)


if __name__ == '__main__':
    main()
